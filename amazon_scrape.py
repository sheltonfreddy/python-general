''' Save the price of a particular product in amazon on a daily basis to an xls file'''


import bs4
import requests
import datetime
import os
import openpyxl
import schedule
import time

url = 'https://www.amazon.com/Apple-MacBook-Display-Graphics-MPXT2LL/dp/B072QGNFMC/ref=sr_1_4?ie=UTF8&qid=1525106415&sr=8-4&keywords=macbook+pro+13+inch+2017'


def get_amazon_price(url):
    res = requests.get(url, headers={
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.110 Safari/537.36"
    })
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    elem = soup.select("#priceblock_ourprice")
    if not os.path.exists('amazon_daily_price.xlsx'):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Amazon Price")
        ws.append([elem[0].text, datetime.datetime.now()])
        wb.save('amazon_daily_price.xlsx')
    else:
        wb = openpyxl.load_workbook('amazon_daily_price.xlsx')
        ws = wb['Amazon Price']
        ws.append([elem[0].text, datetime.datetime.now()])
        wb.save('amazon_daily_price.xlsx')


# schedule.every(1).minutes.do(get_amazon_price, url)
schedule.every().day.at("07:00").do(get_amazon_price, url)

while True:
    schedule.run_pending()
    time.sleep(1)
